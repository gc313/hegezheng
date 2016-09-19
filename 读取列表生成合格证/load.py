
from openpyxl import load_workbook
wb = load_workbook(filename ='合格证列表.xlsx')
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name('Sheet1')
data_list = []

'''
表的格式如下

产品名称	部件名称	货号	生产批号	灭菌批号	检验日期	起始编号	结束编号

'''
#print(ws.cell(row = 0, column = 0 ).value)

#把数据存到列表中     


for rx in range(2, 8):
        
    temp_list = []    
    w1 = ws.cell(row = rx,column = 1).value
    w2 = ws.cell(row = rx,column = 2).value
    w3 = ws.cell(row = rx,column = 3).value
    w4 = ws.cell(row = rx,column = 4).value
    w5 = ws.cell(row = rx,column = 5).value
    w6 = ws.cell(row = rx,column = 6).value
    w7 = ws.cell(row = rx,column = 7).value
    w8 = ws.cell(row = rx,column = 8).value
    temp_list = [w1,w2,w3,w4,w5,w6,w7,w8]
       
    data_list.append(temp_list)

#print (data_list)
